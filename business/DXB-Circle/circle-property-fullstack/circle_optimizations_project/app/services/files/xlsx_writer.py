from typing import List, Dict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
def _autosize(ws):
    for col in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(10, length + 2), 40)
def write_allocations_plan_xlsx(path: str, summary: Dict, allocations: List[Dict]):
    wb = Workbook(); ws1 = wb.active; ws1.title = "Summary"; ws1.append(["Metric", "Value"])
    for k, v in summary.items(): ws1.append([k, v])
    _autosize(ws1); ws2 = wb.create_sheet("Allocations")
    if allocations: ws2.append(list(allocations[0].keys())); 
    for r in allocations: ws2.append(list(r.values()))
    _autosize(ws2); wb.save(path); return path
def write_capex_gantt_xlsx(path: str, schedule: List[Dict]):
    wb = Workbook(); ws = wb.active; ws.title = "Capex Schedule"; ws.append(["Month", "Total Spend", "Breakdown"])
    for row in schedule: breakdown = " · ".join(f"{p['project_id']}:{p['spend']:.0f}" for p in row.get("projects", [])); ws.append([row["month"], row["spend"], breakdown])
    _autosize(ws); wb.save(path); return path
def write_debt_amort_xlsx(path: str, tranches: List[Dict], months: int = 12):
    wb = Workbook(); ws = wb.active; ws.title = "Amort (IO Approx)"; ws.append(["Month"] + [t["name"] for t in tranches] + ["Total Interest"])
    for m in range(1, months + 1):
        interests = []; 
        for t in tranches: rate = t.get("rate") or t.get("effective_rate_base") or 0.0; interests.append(t["amount"] * rate / 12.0)
        ws.append([m, *interests, sum(interests)])
    _autosize(ws); wb.save(path); return path
def write_leasing_offer_xlsx(path: str, mix: List[Dict], kpis: Dict):
    wb = Workbook(); ws = wb.active; ws.title = "Offer Plan"; ws.append(["Package", "Units", "Share", "WAULT contrib (m)"])
    for m in mix: ws.append([m["package"], m["units"], m["share"], m["wault_contrib"]])
    _autosize(ws); ws2 = wb.create_sheet("KPIs"); ws2.append(["Metric", "Value"])
    for k, v in kpis.items(): ws2.append([k, v])
    _autosize(ws2); wb.save(path); return path
