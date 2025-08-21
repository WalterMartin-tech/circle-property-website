import io
import os

from backend.core.calculations.ipa_engine.engine import Inputs, run_calc
from fastapi import Depends, FastAPI, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from reportlab.pdfgen import canvas

from backend.apps.auth.routes import require_auth
from backend.apps.auth.routes import router as auth_router

# ✅ feature routers
from backend.apps.carmodels.routes import router as carmodels_router
from backend.apps.core.routes import router as core_router
from backend.apps.insurance.routes import router as insurance_router
from backend.apps.preview.routes import router as preview_router

app = FastAPI(title="IPA Calculator API")

# ✅ attach feature routers
app.include_router(carmodels_router, prefix="/carmodels", tags=["Car Models"])
app.include_router(insurance_router, prefix="/insurance", tags=["Insurance"])
app.include_router(preview_router, prefix="/preview", tags=["Preview"])
app.include_router(auth_router, prefix="/auth", tags=["Auth"])
app.include_router(core_router, tags=["Core"])  # 👈 NEW


# ✅ attach dev-only router (conditionally)
if os.getenv("DEV_ALLOW_PUBLIC_COMPUTE") == "1":
    from apps.dev.routes import router as dev_router

    app.include_router(dev_router, tags=["Dev"])


# CORS for local dev
origins = [
    o.strip()
    for o in os.getenv(
        "FRONTEND_ORIGINS", "http://localhost:3000,http://127.0.0.1:3000"
    ).split(",")
    if o.strip()
]
koyeb = os.getenv("KOYEB_APP_DOMAIN")
if koyeb:
    origins.append(f"https://{koyeb}")
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/calculate")
async def calculate(req: Request, _=Depends(require_auth)):
    body = await req.json()
    try:
        inp = Inputs(
            principal=float(
                body.get("principal", 0) or body.get("Inputs__principal", 0)
            ),
            rate=float(body.get("rate", 0) or body.get("Inputs__rate", 0)),
            term_months=int(
                body.get("term_months", 0) or body.get("Inputs__term_months", 0)
            ),
            balloon=float(body.get("balloon", 0) or body.get("Inputs__balloon", 0)),
            vat_rate=float(
                body.get("vat_rate", 0.18) or body.get("Inputs__vat_rate", 0.18)
            ),
            asset_vat=float(
                body.get("asset_vat", 0) or body.get("Inputs__asset_vat", 0)
            ),
            telematics_monthly=float(
                body.get("telematics_monthly", 0)
                or body.get("Inputs__telematics_monthly", 0)
            ),
            include_irc=bool(
                body.get("include_irc", True)
                if body.get("include_irc", True) is not None
                else True
            ),
            include_banking=bool(
                body.get("include_banking", True)
                if body.get("include_banking", True) is not None
                else True
            ),
        )
    except Exception:
        import traceback

        print("⚠️ Failed to parse Inputs:", traceback.format_exc())
        inp = body

    res = run_calc(inp)
    # Ensure a totals dict is present for tests/clients
    if not isinstance(res.get("totals"), dict):
        t = {}
        for k in ("annuity", "ipa_vat", "asset_vat", "vat_delta"):
            if k in res:
                t[k] = res[k]
        res["totals"] = t
    return res


@app.post("/export/xlsx")
async def export_xlsx(req: Request, _=Depends(require_auth)):
    data = await req.json()
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculation"
    ws["A1"] = "Example Metric"
    ws["B1"] = "Another"
    ws["A2"] = 123.45
    ws["B2"] = 67.89
    ws["A4"] = "Payload Echo"
    ws["B4"] = str(data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="calculation.xlsx"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/export/pdf")
async def export_pdf(req: Request, _=Depends(require_auth)):
    data = await req.json()
    buf = io.BytesIO()
    c = canvas.Canvas(buf)
    c.setFont("Helvetica", 12)
    c.drawString(72, 800, "IPA Calculator - Demo PDF")
    c.drawString(72, 780, "Example Metric: 123.45")
    c.drawString(72, 760, "Another: 67.89")
    c.drawString(72, 740, f"Payload: {str(data)[:80]}...")
    c.showPage()
    c.save()
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="calculation.pdf"'}
    return StreamingResponse(buf, media_type="application/pdf", headers=headers)
