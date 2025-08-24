import io

from backend.apps.auth.routes import require_auth
from backend.core.calculations.equilibrium import (
    solve_equilibrium_f,
    solve_equilibrium_f_bisect,
    solve_equilibrium_principal,
)
from backend.core.calculations.ipa_engine.engine import Inputs, run_calc
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from reportlab.pdfgen import canvas

router = APIRouter()


# --- Calculate ---
@router.post("/calculate")
async def calculate(inp: Inputs, _=Depends(require_auth)):
    try:
        res = run_calc(inp)
    except Exception as e:
        msg = str(e)
        if "missing" in msg.lower():
            raise HTTPException(status_code=400, detail="Missing required field")
        raise HTTPException(status_code=400, detail=msg)

    # Backward-compatibility: ensure "totals" dict exists
    if not isinstance(res.get("totals"), dict):
        totals = {}
        for k in ("annuity", "ipa_vat", "asset_vat", "vat_delta"):
            if k in res:
                totals[k] = res[k]
        res["totals"] = totals

    return res


# --- Equilibrium: principal ---
@router.post("/equilibrium/principal")
async def equilibrium_principal(inp: Inputs, _=Depends(require_auth)):
    try:
        return solve_equilibrium_principal(inp.model_dump())
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# --- Equilibrium: f ---
@router.post("/equilibrium/f")
async def equilibrium_f(inp: Inputs, _=Depends(require_auth)):
    try:
        return solve_equilibrium_f(inp.model_dump())
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# --- Equilibrium: f_bisect ---
@router.post("/equilibrium/f_bisect")
async def equilibrium_f_bisect(inp: Inputs, _=Depends(require_auth)):
    try:
        return solve_equilibrium_f_bisect(inp.model_dump())
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# --- Export XLSX ---
@router.post("/export/xlsx")
async def export_xlsx(inp: Inputs, _=Depends(require_auth)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculation"
    ws["A1"] = "Example Metric"
    ws["B1"] = "Another"
    ws["A2"] = 123.45
    ws["B2"] = 67.89
    ws["A4"] = "Payload Echo"
    ws["B4"] = str(inp.model_dump())

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="calculation.xlsx"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# --- Export PDF ---
@router.post("/export/pdf")
async def export_pdf(inp: Inputs, _=Depends(require_auth)):
    buf = io.BytesIO()
    c = canvas.Canvas(buf)
    c.setFont("Helvetica", 12)
    c.drawString(72, 800, "IPA Calculator - Demo PDF")
    c.drawString(72, 780, "Example Metric: 123.45")
    c.drawString(72, 760, "Another: 67.89")
    c.drawString(72, 740, f"Payload: {str(inp.model_dump())[:80]}...")
    c.showPage()
    c.save()
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="calculation.pdf"'}
    return StreamingResponse(buf, media_type="application/pdf", headers=headers)
